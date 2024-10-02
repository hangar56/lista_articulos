import openpyxl
import argparse
import sys

def extract_hyperlinks(xlsx_file, txt_file):
    # Load the workbook
    wb = openpyxl.load_workbook(xlsx_file)
    
    # Open the text file for writing
    with open(txt_file, 'w') as f:
        # Iterate through all sheets
        for sheet in wb.sheetnames:
            ws = wb[sheet]
            
            # Iterate through all cells in the sheet
            for row in ws.iter_rows():
                for cell in row:
                    # Check if the cell has a hyperlink
                    if cell.hyperlink:
                        # Write the hyperlink target to the text file
                        f.write(f"{cell.hyperlink.target}\n")

    print(f"Hyperlinks have been extracted to {txt_file}")

def main():
    # Set up argument parser
    parser = argparse.ArgumentParser(description='Extract hyperlinks from an XLSX file to a TXT file.')
    parser.add_argument('input', help='Input XLSX file')
    parser.add_argument('output', help='Output TXT file')

    # Parse arguments
    args = parser.parse_args()

    # Call the extraction function with provided arguments
    extract_hyperlinks(args.input, args.output)

if __name__ == "__main__":
    main()